import pandas as pd
import os
from openpyxl import load_workbook

def append_to_excel(df, target_path):
    """
    Appends a dataframe to an existing Excel file or creates a new one.
    """
    if not os.path.exists(target_path):
        df.to_excel(target_path, index=False)
        return f"Created new file: {target_path}"
    
    try:
        with pd.ExcelWriter(target_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            # Load existing sheet or create new one
            try:
                existing_df = pd.read_excel(target_path)
                combined_df = pd.concat([existing_df, df], ignore_index=True)
                combined_df.to_excel(writer, index=False, sheet_name='Sheet1')
            except Exception:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # Apply formatting using openpyxl directly
        wb = load_workbook(target_path)
        ws = wb.active
        # Find 'Total Value' column index
        header = [cell.value for cell in ws[1]]
        if "Total Value" in header:
            col_idx = header.index("Total Value") + 1
            # Apply format to all rows in this column (except header)
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = '#,##0.00'
        wb.save(target_path)
        
        return f"Appended data to: {target_path}"
    except Exception as e:
        return f"Error appending to Excel: {e}"

def load_master_data(master_path):
    if os.path.exists(master_path):
        return pd.read_excel(master_path)
    return None
